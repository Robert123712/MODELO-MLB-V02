# =============================================================
# GENERADOR DE EXCEL — Predicciones MLB con formato profesional
# Uso:  python -u generar_excel.py [fecha]
# Ej:   python -u generar_excel.py 07/02/2026
# =============================================================

import sys, os
sys.stdout.reconfigure(encoding="utf-8")

from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

import modelo_diario as m

# ── Estilos ──
AZUL = "1F4E79"
AZUL_CLARO = "D6E4F0"
VERDE = "C6EFCE"
ROJO = "FFC7CE"
AMARILLO = "FFEB9C"
GRIS = "F2F2F2"
BLANCO = "FFFFFF"

header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor=AZUL)
subheader_font = Font(name="Calibri", bold=True, color=AZUL, size=10)
data_font = Font(name="Calibri", size=10)
bold_font = Font(name="Calibri", bold=True, size=10)
title_font = Font(name="Calibri", bold=True, color=AZUL, size=14)
pct_fmt = '0.0%'
dec_fmt = '0.00'
thin_border = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
fill_par = PatternFill("solid", fgColor=GRIS)

# La conversion a momios vive en modelo_diario (reglas del mercado: EVEN en vez
# de -100, vig cargado al no favorito, escalones reales). Aqui solo se reusa.
prob_a_momio = m.prob_a_momio
momio_mercado = m.momio_mercado

def estilo_celda(ws, row, col, value, font=data_font, fill=None, fmt=None, align=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    cell.border = thin_border
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    if align:
        cell.alignment = align
    return cell

def escribir_encabezados(ws, row, headers, fill=header_fill, font=header_font):
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = font
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def colorear_overs(ws, row, col, p):
    cell = ws.cell(row=row, column=col)
    if p >= 0.65:
        cell.fill = PatternFill("solid", fgColor=VERDE)
    elif p <= 0.35:
        cell.fill = PatternFill("solid", fgColor=ROJO)
    else:
        cell.fill = PatternFill("solid", fgColor=AMARILLO)

# ══════════════════════════════════════════════════════════════

def generar(fecha=None):
    hoy = fecha or date.today().strftime("%m/%d/%Y")

    print(f"📡 Obteniendo datos para {hoy}...", flush=True)
    juegos = m.statsapi.schedule(date=hoy)
    modelables = [j for j in juegos if j["status"] in ("Scheduled", "Pre-Game", "Warmup")
                  and j["away_probable_pitcher"] and j["home_probable_pitcher"]]

    if not modelables:
        print("⚠ No hay juegos modelables para esta fecha.", flush=True)
        return

    print(f"✅ {len(modelables)} juegos encontrados. Procesando...", flush=True)

    _frac_f5 = m.f5_frac_liga(hoy)

    # Se evalua CADA JUEGO UNA SOLA VEZ. Antes cada hoja repetia la tuberia
    # completa: el mismo juego se simulaba 3 veces (3x llamadas y 3x Monte Carlo).
    evaluados = [r for r in (m.evaluar_juego(j, hoy, _frac_f5) for j in modelables) if r]
    if not evaluados:
        print("⚠ Ningun juego pudo evaluarse.", flush=True)
        return

    wb = openpyxl.Workbook()
    # ── SHEET 1: RESUMEN ──
    ws1 = wb.active
    ws1.title = "Resumen"
    ws1.sheet_properties.tabColor = AZUL

    ws1.merge_cells("A1:N1")
    titulo = ws1.cell(row=1, column=1, value=f"⚾ PREDICCIONES MLB — {hoy}")
    titulo.font = title_font
    ws1.row_dimensions[1].height = 30

    ws1.merge_cells("A2:N2")
    ws1.cell(row=2, column=1, value=(
        f"Calibración: amortigua={m.AMORTIGUA} | dispersion_k={m.DISPERSION_K} | "
        f"base={m.AJUSTE_BASE} | F5 frac={_frac_f5:.3f} | Sims={m.N_SIMS:,}"
    )).font = Font(name="Calibri", italic=True, color="666666", size=9)

    fila = 4
    headers = [
        "Visitante", "Casa", "Abridor V", "Abridor C",
        "FIP V", "FIP C", "IP V", "IP C",
        "Bullpen V", "Bullpen C",
        "RG V", "RG C", "Split V", "Split C", "Park",
        "DEF V", "DEF C",
        "λ V", "λ C", "Total λ",
        "ML V %", "ML V Justo", "ML V Casa", "ML C %", "ML C Justo", "ML C Casa",
        "RL V +1.5", "RL C -1.5",
        "NRFI",
        "O5.5", "O6.5", "O7.5", "O8.5", "O9.5", "O10.5",
    ]
    escribir_encabezados(ws1, fila, headers)
    ws1.row_dimensions[fila].height = 35

    data_fila = fila + 1
    COL_OVERS = 30          # primera columna de overs (1-indexada)
    COL_ML_PCT = (21, 24)
    COL_ML_MOM = (22, 23, 25, 26)

    for idx, r in enumerate(evaluados):
        overs, p_casa = r["overs"], r["p_casa"]
        datos = [
            r["visita"], r["casa"], r["abridor_v"], r["abridor_c"],
            round(r["fip_v"], 2), round(r["fip_c"], 2), round(r["ip_v"], 1), round(r["ip_c"], 1),
            round(r["bp_v"]["fip"], 2), round(r["bp_c"]["fip"], 2),
            round(r["rg_v"], 2), round(r["rg_c"], 2),
            round(r["split_v"], 3), round(r["split_c"], 3), r["park"],
            round(r["def_v"], 3), round(r["def_c"], 3),
            round(r["lam_v"], 2), round(r["lam_c"], 2), round(r["total_esp"], 2),
            round(r["p_visita"], 4), prob_a_momio(r["p_visita"]), momio_mercado(r["p_visita"]),
            round(p_casa, 4), prob_a_momio(p_casa), momio_mercado(p_casa),
            round(r["p_visita_rl"], 4), round(r["p_casa_rl"], 4),
            round(r["nrfi"]["nrfi"], 4),
            round(overs[5.5], 4), round(overs[6.5], 4), round(overs[7.5], 4),
            round(overs[8.5], 4), round(overs[9.5], 4), round(overs[10.5], 4),
        ]
        row = data_fila + idx
        alt_fill = PatternFill("solid", fgColor=GRIS) if idx % 2 == 0 else None
        for col, val in enumerate(datos, 1):
            cell = ws1.cell(row=row, column=col, value=val)
            cell.font = data_font
            cell.border = thin_border
            if alt_fill:
                cell.fill = alt_fill
            if isinstance(val, float) and col >= COL_OVERS:
                cell.number_format = pct_fmt
                colorear_overs(ws1, row, col, val)
            elif isinstance(val, float):
                cell.number_format = dec_fmt
            if col in COL_ML_PCT or col == 29:
                cell.number_format = pct_fmt
            if col in COL_ML_MOM:
                cell.alignment = Alignment(horizontal="center")

    # Anchos de columna
    anchos = [20, 20, 18, 18, 7, 7, 6, 6, 9, 9, 7, 7, 7, 7, 5,
              7, 7, 7, 7, 8, 9, 9, 9, 9, 9, 9, 9, 9, 7, 6, 6, 6, 6, 6, 6]
    for i, a in enumerate(anchos, 1):
        ws1.column_dimensions[get_column_letter(i)].width = a
    ws1.freeze_panes = ws1.cell(row=fila + 1, column=1)

    # Total slate
    if evaluados:
        prom = sum(r['total_esp'] for r in evaluados) / len(evaluados)
        fin_fila = data_fila + len(evaluados)
        ws1.merge_cells(f"A{fin_fila}:G{fin_fila}")
        cel = ws1.cell(row=fin_fila, column=1, value=f"📊 Promedio total del slate: {prom:.2f} carreras")
        cel.font = Font(name="Calibri", bold=True, italic=True, size=10)

    # ── SHEET 2: DETALLE X JUEGO ──
    ws2 = wb.create_sheet("Detalle por Juego")
    ws2.sheet_properties.tabColor = "2E75B6"

    fila2 = 1
    for r in evaluados:
        visita, casa, f5 = r["visita"], r["casa"], r["f5"]
        overs, p_casa = r["overs"], r["p_casa"]
        p_visita = r["p_visita"]

        # ── CABECERA DEL JUEGO ──
        ws2.merge_cells(f"A{fila2}:J{fila2}")
        cell = ws2.cell(row=fila2, column=1,
                        value=f"{visita} @ {casa}  |  {r['abridor_v']} vs {r['abridor_c']}"
                              + ("   ⚠ abridor sin stats: estimado" if r["estimado"] else ""))
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill("solid", fgColor=AZUL)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws2.row_dimensions[fila2].height = 28
        fila2 += 1

        escribir_encabezados(ws2, fila2, ["Métrica", "Visitante", "Casa"],
                             fill=PatternFill("solid", fgColor=AZUL_CLARO), font=subheader_font)
        fila2 += 1

        def par(fila, label, val_v, val_c, fmt_override=None):
            celdas = [estilo_celda(ws2, fila, 1, label, font=bold_font),
                      estilo_celda(ws2, fila, 2, val_v, fmt=fmt_override),
                      estilo_celda(ws2, fila, 3, val_c, fmt=fmt_override)]
            if fila % 2 == 0:
                for c in celdas:
                    c.fill = fill_par

        par(fila2, "FIP", round(r["fip_v"], 2), round(r["fip_c"], 2)); fila2 += 1
        par(fila2, "IP Esperadas", round(r["ip_v"], 1), round(r["ip_c"], 1)); fila2 += 1
        par(fila2, "Mano", r["mano_v"] or "?", r["mano_c"] or "?"); fila2 += 1
        par(fila2, "Bullpen FIP", round(r["bp_v"]["fip"], 2), round(r["bp_c"]["fip"], 2)); fila2 += 1
        par(fila2, "R/G (park-ajustada)", round(r["rg_v"], 2), round(r["rg_c"], 2)); fila2 += 1
        par(fila2, "Split vs Mano", round(r["split_v"], 3), round(r["split_c"], 3)); fila2 += 1
        par(fila2, "Factor DEF", round(r["def_v"], 3), round(r["def_c"], 3)); fila2 += 1
        par(fila2, "Carreras Esperadas (λ)", round(r["lam_v"], 2), round(r["lam_c"], 2)); fila2 += 1
        par(fila2, "Total Esperado", round(r["total_esp"], 2), "—"); fila2 += 1
        par(fila2, "Park Factor", r["park"], "—"); fila2 += 1
        par(fila2, "Moneyline Prob", round(p_visita, 4), round(p_casa, 4), pct_fmt); fila2 += 1
        par(fila2, "Moneyline Momio justo", prob_a_momio(p_visita), prob_a_momio(p_casa)); fila2 += 1
        par(fila2, "Moneyline línea casa", momio_mercado(p_visita), momio_mercado(p_casa)); fila2 += 1
        par(fila2, "Run Line +1.5 / -1.5", round(r["p_visita_rl"], 4), round(r["p_casa_rl"], 4), pct_fmt); fila2 += 1
        par(fila2, "Total del equipo O4.5", round(r["tt_visita"][4.5], 4), round(r["tt_casa"][4.5], 4), pct_fmt); fila2 += 1
        fila2 += 1

        # Overs / Unders
        LINEAS_XL = [5.5, 6.5, 7.5, 8.5, 9.5, 10.5, 11.5, 12.5]
        escribir_encabezados(ws2, fila2, [""] + [f"O{ln}" for ln in LINEAS_XL],
                             fill=PatternFill("solid", fgColor=AZUL_CLARO), font=subheader_font)
        fila2 += 1
        for etiqueta, transform in (("Over", lambda p: p), ("Under", lambda p: 1 - p)):
            estilo_celda(ws2, fila2, 1, etiqueta, font=bold_font)
            for ci, ln in enumerate(LINEAS_XL):
                val = transform(overs[ln])
                estilo_celda(ws2, fila2, ci + 2, round(val, 4), fmt=pct_fmt)
                colorear_overs(ws2, fila2, ci + 2, val)
            fila2 += 1
        fila2 += 1

        # F5 + NRFI
        ws2.cell(row=fila2, column=1, value="Primeras 5 Entradas (F5) y 1ª entrada").font = Font(
            name="Calibri", bold=True, color="2E75B6", size=11)
        fila2 += 1
        escribir_encabezados(ws2, fila2, ["Métrica", "Valor"],
                             fill=PatternFill("solid", fgColor=AZUL_CLARO), font=subheader_font)
        fila2 += 1
        detalle = [
            ("Carreras λ V (F5)", round(f5["lam_v"], 2)),
            ("Carreras λ C (F5)", round(f5["lam_c"], 2)),
            ("Total F5", round(f5["total_esp"], 2)),
            ("ML F5 Casa", f"{f5['p_casa']:.1%} (justo {prob_a_momio(f5['p_casa'])} / casa {momio_mercado(f5['p_casa'])})"),
            ("ML F5 Visita", f"{f5['p_visita']:.1%} (justo {prob_a_momio(f5['p_visita'])} / casa {momio_mercado(f5['p_visita'])})"),
            ("ML F5 Empate", f"{f5['p_empate']:.1%} (justo {prob_a_momio(f5['p_empate'])} / casa {momio_mercado(f5['p_empate'])})"),
            ("RL F5 Casa +0.5", f"{f5['rl_casa']:.1%}"),
            ("RL F5 Visita +0.5", f"{f5['rl_visita']:.1%}"),
            ("Total F5 O4.5", f"{f5['overs'][4.5]:.1%}"),
            ("NRFI (1ª sin carreras)", f"{r['nrfi']['nrfi']:.1%} (justo {prob_a_momio(r['nrfi']['nrfi'])} / casa {momio_mercado(r['nrfi']['nrfi'])})"),
            ("YRFI", f"{r['nrfi']['yrfi']:.1%} (justo {prob_a_momio(r['nrfi']['yrfi'])} / casa {momio_mercado(r['nrfi']['yrfi'])})"),
        ]
        for lbl, val in detalle:
            estilo_celda(ws2, fila2, 1, lbl, font=bold_font)
            estilo_celda(ws2, fila2, 2, val)
            if fila2 % 2 == 0:
                for c in (1, 2):
                    ws2.cell(row=fila2, column=c).fill = fill_par
            fila2 += 1
        fila2 += 2

    # Anchos sheet 2
    for c in range(1, 11):
        ws2.column_dimensions[get_column_letter(c)].width = 18
    ws2.column_dimensions["A"].width = 22

    # ── SHEET 3: JUGADAS DE VALOR (+EV) ──
    ws3 = wb.create_sheet("Jugadas +EV")
    ws3.sheet_properties.tabColor = "00B050"

    ws3.merge_cells("A1:E1")
    ws3.cell(row=1, column=1, value="💰 JUGADAS CON VALOR ESPERADO POSITIVO").font = title_font
    ws3.merge_cells("A2:E2")
    ws3.cell(row=2, column=1, value="(Requiere ODDS_API_KEY activada en variable de entorno)").font = Font(name="Calibri", italic=True, color="999999", size=9)

    headers3 = ["Juego", "Mercado", "Pick", "Prob Modelo", "Momio", "EV", "Casa"]
    escribir_encabezados(ws3, 4, headers3)

    f3 = 5
    ods = m.valor.obtener_odds()
    for r in evaluados:
        jugadas = m.valor.analizar_juego(m.valor.buscar(ods, r["visita"], r["casa"]),
                                         r["visita"], r["casa"], r["p_casa"], r["overs"])
        for jg in jugadas:
            estilo_celda(ws3, f3, 1, f"{r['visita']} @ {r['casa']}", font=data_font)
            estilo_celda(ws3, f3, 2, jg["mercado"], font=data_font)
            estilo_celda(ws3, f3, 3, jg["pick"], font=bold_font)
            estilo_celda(ws3, f3, 4, round(jg["p_modelo"], 4), fmt=pct_fmt, font=data_font)
            mom = jg["momio"]
            estilo_celda(ws3, f3, 5, f"{mom:+d}" if mom > 0 else str(mom), font=data_font)
            estilo_celda(ws3, f3, 6, round(jg["ev"], 4), fmt=pct_fmt,
                         fill=PatternFill("solid", fgColor=VERDE))
            estilo_celda(ws3, f3, 7, jg.get("libro", ""), font=data_font)
            f3 += 1

    if f3 == 5:
        ws3.merge_cells(f"A3:F3")
        ws3.cell(row=3, column=1, value="⚠ Sin ODDS_API_KEY — no hay datos de mercado para detectar valor").font = Font(name="Calibri", italic=True, color="FF0000", size=10)

    for c in range(1, 8):
        ws3.column_dimensions[get_column_letter(c)].width = 20

    # ── Guardar ──
    os.makedirs("exports", exist_ok=True)
    filename = f"exports/MLB_Predicciones_{hoy.replace('/', '-')}.xlsx"
    wb.save(filename)
    print(f"\n✅ Excel generado: {os.path.abspath(filename)}", flush=True)
    print(f"   {len(evaluados)} juegos modelados", flush=True)
    print(f"   Abrelo en Excel para ver formato completo con colores", flush=True)

if __name__ == "__main__":
    fecha = sys.argv[1] if len(sys.argv) > 1 else None
    generar(fecha)
